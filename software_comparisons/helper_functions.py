# -*- coding: utf-8 -*-
"""
Created on Wed Oct 20 01:12:16 2021

@author: User
"""
import pickle
import openpyxl
import os
import pandas as pd
import xlrd
import collections
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from termcolor import colored

# COLLECTING USER INPUT ON GENE OF INTEREST
def collect_input():
    gene_name = input("Enter name of target gene (only letters): ")
    mRNA_sequence = input("Enter nucleotide sequence: ")
    FASTA_sequence = input("Enter sequence in FASTA format: ")
    email = input("Enter your email address (for siRNA batch job submission): ")
    return gene_name, mRNA_sequence, FASTA_sequence, email

# GENERATING CORRECT FORMATTING OF EXCEL

def generate_workbook(gene_name, mRNA_sequence, FASTA_sequence, filename="test.xlsx"):
    excel_workbook = Workbook()
    sheet = excel_workbook.active

    sheet['A1'] = 'name of gene'
    sheet['A2'] = 'mRNA nucleotide sequence'
    sheet['A3'] = 'mRNA FASTA sequence'
    sheet['B1'] = gene_name
    sheet['B2'] = mRNA_sequence
    sheet['B3'] = FASTA_sequence

    program_indices = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P', 'R']
    program_names = ['siRNA_wizard', 'sFold_over12', 'siDESIGN_center', 'BLOCKIT', 'Eurofins_siMax', 'RNAiExplorer_Genelink', 'Oligowalk', 'siDirect', 'IDT']
    titles_row = '4'

    for i in range(len(program_names)):
        cellname = program_indices[i] + titles_row
        sheet[cellname] = program_names[i]

    excel_workbook.save(filename)
    
    return excel_workbook, sheet
    
def read_workbook(file_path=r'C:\Users\User\Desktop\ETH_NSC\Yanik_lab\siRNADesign-master\software_comparisons\test.xlsx'):
    excel_workbook = load_workbook(file_path, data_only=True)
    sheet = excel_workbook.active
    
    return excel_workbook, sheet

"""
def save_workbook(excel_workbook,filename='test.xlsx'):
    excel_workbook.save(filename)
    
    return excel_workbook
"""

# SAVE LIST OF TARGET POSITIONS TO DISK
def pickle_positions(filename, targetpositions_list):
    # filename e.g. "wiz_list", targetpositions_list e.g. targetpositions_wiz
    open_file = open(filename,"wb")
    pickle.dump(targetpositions_list, open_file)
    open_file.close()
    
def open_pickle(filename):
    open_file = open(filename, "rb")
    loaded_list = pickle.load(open_file)
    open_file.close()
    
    return loaded_list

# FIND LOCATION OF siRNA_WIZARD HEADER IN EXCEL
def find_title_location(target_positions, excel_workbook, sheet, cell_title):
    # cell_titles: 'siRNA_wizard', 'sFold_over12', 'siDESIGN_center', 'BLOCKIT', 'IDT', 'Eurofins_siMax', 'RNAiExplorer_Genelink', 'Oligowalk', 'siDirect'
    
    for row in sheet.iter_rows():
        for cell in row:
            #print(cell.value)
            #print(cell.coordinate)
            if cell.value == cell_title:
                title_position = cell.coordinate

    col_row = coordinate_from_string(title_position)
    #col = column_index_from_string(col_row[0])
    #row = col_row[1]

    return title_position, col_row

# WRITE TARGET POSITION UNDER HEADER IN EXCEL        
def add_results(target_positions, excel_workbook, sheet, col_row):
    
    for i in range(len(target_positions)):
        new_row = col_row[1] + (i+1)

        #new_row = str(col_row[1] + (i+1))
        #coord_tuple = (col_row[0], new_row)
        #new_coord = coord_tuple[0] + coord_tuple[1]
        #sheet[new_coord] = target_positions[i]
        
        sheet.cell(row=new_row, column=column_index_from_string(col_row[0])).value = target_positions[i]
        
        excel_workbook.save('test.xlsx')
        
    return sheet

def multiple_recommended(filename, workbook, sheet):
    allstartpositions = []
    
    for row in sheet.iter_rows():
        for cell in row:
            allstartpositions.append(cell.value)
            
    # Remove all none  values
    allstartpositions = [i for i in allstartpositions if i]
    # Remove header info
    allstartpositions = allstartpositions[13:] 
    
    # Get duplicates
    duplicate_list = [item for item, count in collections.Counter(allstartpositions).items() if count > 1]
    triplicate_list = [item for item, count in collections.Counter(allstartpositions).items() if count > 2]
    
    return duplicate_list, triplicate_list

# HIGHLIGHT SEQUENCES TARGETED 3 TIMES

def colour_targets(sequence,sense_list):
        
    # remove TT overhangs
    list_noTT = []
    for sense in sense_list:
        list_noTT.append(sense[:-2])
           
    # convert uracils to thymines, to match the NT sequence (despite input being mRNA, ncbi still denotes with Ts)
    list_U2T = []
    for sense in list_noTT:
        sense_T = sense.replace('U','T')
        list_U2T.append(sense_T)
        
    # create copy of sequence in which we will highlight homologous regions
    colour_sequence = sequence
    
    # locate each match within the sequence, format it to be green and bold
    for i in range(len(list_U2T)):
        sense = list_U2T[i]
        print(colored(sense,'green', attrs=['bold']))
        colour_sequence = colour_sequence.replace(sense,colored(sense,'green',attrs=['bold']))
    
    return colour_sequence
       


    
"""
# FIND DUPLICATE EXCEL ENTRIES
def find_duplicates(excel_workbook, sheet):
    
    for row in sheet.iter_rows():
        for cell in row:
            value1 = cell.value
            coordinate1 = cell.coordinate
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.coordinate == coordinate1:
                        print("identity")
                    else:
                        if cell.value == value1:
                            print("A duplicate was found")
                            print(cell.value)
                            print(value1)
                        else:
                            print("not a duplicate")
"""
                
            
            

# -*- coding: utf-8 -*-
"""
Created on Mon Jul 18 09:49:46 2022

@author: User
"""
from scipy import stats
from scipy.stats import pearsonr
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook


def plot_correlation(inhibition,sums,title='Correlation plot',xlabel='% inhibition',ylabel='score'):
    print(title)
    
    # plot sums (y-axis) over inhibition (x-axis)
    plt.figure
    plt.plot(inhibition,sums,color="black")
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    
    # get Pearson's correlation coefficient
    corr,_ = pearsonr(inhibition,sums)
    print('Pearsons correlation: %.3f' % corr)
    

def add_to_excel(excel_workbook,sheet,title,column,data,filename='Huesken_Corr.xlsx'):
    sheet[column + '1'] = title
    
    for i in range(len(data)):
        cellname = column + str(i+2)
        sheet[cellname] = data[i]
    excel_workbook.save(filename)
    
    return excel_workbook, sheet

##################################################
# CREATE AN EXCEL TO STORE RESULTS
##################################################

def excel_results(filename):
    excel_workbook = Workbook()
    sheet = excel_workbook.active
    
    sheet['A2'] = 'Parameter'
    sheet['B2'] = 'Meaning'
    
    sheet['A3'] = '6'
    sheet['B3'] = 'GC content'
    
    sheet['A4'] = '7'
    sheet['B4'] = 'Asymm'
    
    sheet['A5'] = '8'
    sheet['B5'] = 'Energy valley'
    
    sheet['A6'] = '9'
    sheet['B6'] = 'GC repeats'
    
    sheet['A7'] = '10'
    sheet['B7'] = 'AT repeats'
    
    sheet['A8'] = '12'
    sheet['B8'] = 'TT overhang'
    
    sheet['A9'] = '13'
    sheet['B9'] = 'A/U at 5-AS'
    
    sheet['A10'] = '14'
    sheet['B10'] = 'G/C at 5-SS'
    
    sheet['A11'] = '15'
    sheet['B11'] = 'A,6,AS'
    
    sheet['A12'] = '16'
    sheet['B12'] = 'A,3,SS'
    
    sheet['A13'] = '17'
    sheet['B13'] = 'A,19,SS'
    
    sheet['A14'] = '18'
    sheet['B14'] = 'no G/C,19,SS'
    
    sheet['A15'] = '19'
    sheet['B15'] = 'no G,13,SS'
    
    sheet['A16'] = '20'
    sheet['B16'] = 'U,10,SS'
    
    sheet['A17'] = 'total score'
    sheet['B17'] = 'sum of params'
    
    excel_workbook.save(filename)

    return excel_workbook, sheet

def excel_addnewdataset(startpos, excel_workbook, sheet, filename):
    
    # startpos = 3 
    
    sheet.cell(2,startpos).value = 'lower avg'
    sheet.cell(2,startpos+1).value = 'higher avg'
    sheet.cell(2,startpos+2).value = 'ttest result'
    sheet.cell(2,startpos+3).value = 'significance'
    
    excel_workbook.save(filename)

    return excel_workbook, sheet

def excel_addaverages(column, excel_workbook, sheet, filename, score6_lowest, score7_lowest, score8_lowest, score9_lowest, score10_lowest, score12_lowest, score13_lowest, score14_lowest, score15_lowest, score16_lowest, score17_lowest, score18_lowest, score19_lowest, score20_lowest):
    sheet.cell(3,column).value = sum(score6_lowest)/len(score6_lowest)
    sheet.cell(4,column).value = sum(score7_lowest)/len(score7_lowest)
    sheet.cell(5,column).value = sum(score8_lowest)/len(score8_lowest)
    sheet.cell(6,column).value = sum(score9_lowest)/len(score9_lowest)
    sheet.cell(7,column).value = sum(score10_lowest)/len(score10_lowest)
    sheet.cell(8,column).value = sum(score12_lowest)/len(score12_lowest)
    sheet.cell(9,column).value = sum(score13_lowest)/len(score13_lowest)
    sheet.cell(10,column).value = sum(score14_lowest)/len(score14_lowest)
    sheet.cell(11,column).value = sum(score15_lowest)/len(score15_lowest)
    sheet.cell(12,column).value = sum(score16_lowest)/len(score16_lowest)
    sheet.cell(13,column).value = sum(score17_lowest)/len(score17_lowest)
    sheet.cell(14,column).value = sum(score18_lowest)/len(score18_lowest)
    sheet.cell(15,column).value = sum(score19_lowest)/len(score19_lowest)
    sheet.cell(16,column).value = sum(score20_lowest)/len(score20_lowest)
    
    excel_workbook.save('sirna_datasets_results.xlsx')
    
def ttests_parameters(column_result, column_significance, pvalue, sheet, score6_lowest, score7_lowest, score8_lowest, score9_lowest, score10_lowest, score12_lowest, score13_lowest, score14_lowest, score15_lowest, score16_lowest, score17_lowest, score18_lowest, score19_lowest, score20_lowest, lowest_sums, score6_highest, score7_highest, score8_highest, score9_highest, score10_highest, score12_highest, score13_highest, score14_highest, score15_highest, score16_highest, score17_highest, score18_highest, score19_highest, score20_highest, highest_sums):
    result6 = stats.ttest_ind(score6_lowest, score6_highest)
    sheet.cell(3,column_result).value = result6[1] # returns p-value
    if result6[1] < pvalue:
        sheet.cell(3,column_significance).value = 'significant'
    else:
        sheet.cell(3,column_significance).value = 'n.s.'
    
    result7 = stats.ttest_ind(score7_lowest, score7_highest)
    sheet.cell(4,column_result).value = result7[1]
    if result7[1] < pvalue:
        sheet.cell(4,column_significance).value = 'significant'
    else:
        sheet.cell(4,column_significance).value = 'n.s.'
        
    
    result8 = stats.ttest_ind(score8_lowest, score8_highest)
    sheet.cell(5, column_result).value = result8[1]
    if result8[1] < pvalue:
        sheet.cell(5,column_significance).value = 'significant'
    else:
        sheet.cell(5,column_significance).value = 'n.s.'
    
    result9 = stats.ttest_ind(score9_lowest, score9_highest)
    sheet.cell(6, column_result).value = result9[1]
    if result9[1] < pvalue:
        sheet.cell(6,column_significance).value = 'significant'
    else:
        sheet.cell(6,column_significance).value = 'n.s.'
    
    result10 = stats.ttest_ind(score10_lowest, score10_highest)
    sheet.cell(7, column_result).value = result10[1]
    if result10[1] < pvalue:
        sheet.cell(7,column_significance).value = 'significant'
    else:
        sheet.cell(7,column_significance).value = 'n.s.'
    
    result12 = stats.ttest_ind(score12_lowest, score12_highest)
    sheet.cell(8, column_result).value = result12[1]
    if result12[1] < pvalue:
        sheet.cell(8,column_significance).value = 'significant'
    else:
        sheet.cell(8,column_significance).value = 'n.s.'
    
    result13 = stats.ttest_ind(score13_lowest, score13_highest)
    sheet.cell(9,column_result).value = result13[1]
    if result13[1] < pvalue:
        sheet.cell(9,column_significance).value = 'significant'
    else:
        sheet.cell(9,column_significance).value = 'n.s.'
    
    result14 = stats.ttest_ind(score14_lowest, score14_highest)
    sheet.cell(10,column_result).value = result14[1]
    if result14[1] < pvalue:
        sheet.cell(10,column_significance).value = 'significant'
    else:
        sheet.cell(10,column_significance).value = 'n.s.'
    
    result15 = stats.ttest_ind(score15_lowest, score15_highest)
    sheet.cell(11,column_result).value = result15[1]
    if result15[1] < pvalue:
        sheet.cell(11,column_significance).value = 'significant'
    else:
        sheet.cell(11,column_significance).value = 'n.s.'
    
    result16 = stats.ttest_ind(score16_lowest, score16_highest)
    sheet.cell(12,column_result).value = result16[1]
    if result16[1] < pvalue:
        sheet.cell(12,column_significance).value = 'significant'
    else:
        sheet.cell(12,column_significance).value = 'n.s.'
    
    result17 = stats.ttest_ind(score17_lowest, score17_highest)
    sheet.cell(13,column_result).value = result17[1]
    if result17[1] < pvalue:
        sheet.cell(13,column_significance).value = 'significant'
    else:
        sheet.cell(13,column_significance).value = 'n.s.'
    
    result18 = stats.ttest_ind(score18_lowest, score18_highest)
    sheet.cell(14,column_result).value = result18[1]
    if result18[1] < pvalue:
        sheet.cell(14,column_significance).value = 'significant'
    else:
        sheet.cell(14,column_significance).value = 'n.s.'
    
    result19 = stats.ttest_ind(score19_lowest, score19_highest)
    sheet.cell(15,column_result).value = result19[1]
    if result19[1] < pvalue:
        sheet.cell(15,column_significance).value = 'significant'
    else:
        sheet.cell(15,column_significance).value = 'n.s.'
    
    result20 = stats.ttest_ind(score20_lowest, score20_highest)
    sheet.cell(16,column_result).value = result20[1]
    if result20[1] < pvalue:
        sheet.cell(16,column_significance).value = 'significant'
    else:
        sheet.cell(16,column_significance).value = 'n.s.'
    
    result21 = stats.ttest_ind(lowest_sums, highest_sums)
    sheet.cell(17,column_result).value = result21[1]
    if result21[1] < pvalue:
        sheet.cell(17,column_significance).value = 'significant'
    else:
        sheet.cell(17,column_significance).value = 'n.s.'
    
    
    return sheet
    

##################################################
# HELPER FUNCTIONS MYSIRNA DATASETS (C,D)
##################################################

# remove space at the start of list entries
def remove_spaces(input_list): # input: e.g. antisense_list, inhibition

    # remove unnecessary spaces

    for i in range(len(input_list)):
        if input_list[i][0] == " ":
            input_list[i] = input_list[i][1:]
            
    return input_list

def remove_overhangs(antisense_withoverhang):
    
    antisense_list = []
    for antisense in antisense_withoverhang:
        antisense_list.append(antisense[:-2])
    
    return antisense_list

# count how many inhibition values are below 10 and above 90
# if plot_hist is set to true, a histogram showing the distribution of inhibition values is shown
def count_lower_higher(inhibition,bottom=10, middle=50, upper=90, plot_hist=True,decimal=False): # turn decimal True if inhibition = 0.1 instead of 10 etc.
        
    # create counting variables

    count_below10 = 0
    count_below50 = 0
    count_above90 = 0
    intermediate = 0

    # convert inhibition values to floats

    for i in range(len(inhibition)):
        inhibition[i] = float(inhibition[i])
    
    # create lists to store indices of lower and higher values
    indices_below10 = []
    indices_above90 = []

    # count lower and higher

    for i in range(len(inhibition)):
    
        value = inhibition[i]
        
        if decimal==False:
        
            if value < bottom:
                count_below10 = count_below10 + 1
                indices_below10.append(i)
            if value < upper:
                count_below50 = count_below50 + 1 
            if value > upper:
                count_above90 = count_above90 + 1 
                indices_above90.append(i)
            if value > middle and value < upper-10:
                intermediate = intermediate + 1
        
        elif decimal==True:
            
            if value < (bottom/100):
                count_below10 = count_below10 + 1
                indices_below10.append(i)
            if value < (upper/100):
                count_below50 = count_below50 + 1 
            if value > (upper/100):
                count_above90 = count_above90 + 1 
                indices_above90.append(i)
            if value > (middle/100) and value < ((upper-10)/100):
                intermediate = intermediate + 1
            
    print("There is " + str(count_below10) + " siRNAs achieving inhibition below " + str(bottom))
    print("There is " + str(count_below50) + " siRNAs achieving inhibition below " + str(middle))
    print("There is " + str(count_above90) + " siRNAs achieving inhibition above " + str(upper))
    print("There is " + str(intermediate) + " siRNAs achieving inhibition between " + str(middle) + " and " + str(upper-10))

    if plot_hist == True:
        # Plot distribution of inhibition values
        plt.hist(inhibition) # it looks like this because a lot of values were excluded from the original dataset

    return indices_below10, indices_above90

# store antisense strands with inhibition values below 10 and above 90
def collect_lower_higher(indices_below10, indices_above90, antisense_list):
    
    # collect antisense strands below 10
    antisense_below10 = []

    for index in indices_below10:
        antisense_below10.append(antisense_list[index])

    # collect antisense strands above 90
    antisense_above90 = []

    for index in indices_above90:
        antisense_above90.append(antisense_list[index])
        
    return antisense_below10, antisense_above90

def normalize_list(inputlist):
    normalized_list = []
    for element in inputlist:
        xnorm = (element - min(inputlist))/(max(inputlist) - min(inputlist))
        normalized_list.append(xnorm)
        
    return normalized_list
        
def plot_bar_chart(lowest_sums, highest_sums,title="Dataset X",ylabel="Parameter score",color="black"):
    # get standard errors
    std_error_low = np.std(lowest_sums, ddof=1) / np.sqrt(len(lowest_sums))
    std_error_high = np.std(highest_sums, ddof=1) / np.sqrt(len(highest_sums))
    std_errors = [std_error_low,std_error_high]
    
    # get averages
    avg_low = sum(lowest_sums)/len(lowest_sums)
    avg_high = sum(highest_sums)/len(highest_sums)
    averages = [avg_low,avg_high]
    
    # label bars
    bars = ('Lowest', 'Highest')
    x_pos = np.arange(len(bars))
    
    #define chart 
    fig, ax = plt.subplots()
    
    #create chart
    ax.bar(x=np.arange(len(averages)), #x-coordinates of bars
           height=averages, #height of bars
           yerr=std_errors, #error bar width
           capsize=4, #length of error bar caps
           color="black") 
    
    # add y-axis title
    plt.ylabel(ylabel)
    
    # Create names on the x axis
    plt.xticks(x_pos, bars)
    
    # add chart title
    plt.title(title)
    

